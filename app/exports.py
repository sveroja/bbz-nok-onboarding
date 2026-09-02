"""Export-Generierung: Klassen-LK-Excel, Klassenbuch-Import, DaZ-Import,
Stammdatenblatt-PDF. Nutzt die admin-hochgeladenen Vorlagen aus vorlagen.py.
"""
import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Font
from pypdf import PdfReader, PdfWriter
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

from . import branding, vorlagen
from .stammdatenblatt_layout import (
    CHECKBOX_FIELDS, PAGE_HEIGHT, PAGE_WIDTH, TEXT_FIELDS,
)

PLZ_OK_LABEL = {True: "ja", False: "nein", None: "unklar"}

GESCHLECHT_KURZFORM = {
    "maennlich": "m",
    "weiblich": "w",
    "divers": "d",
    "keine_angabe": "",
}


def _autosize_columns(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 40)


# ---------------------------------------------------------------------------
# Excel fuer die Klassen-LK (wichtigste Daten, kein externes Zielformat)
# ---------------------------------------------------------------------------

def build_klassen_lk_excel(regs) -> io.BytesIO:
    from .models import Bildungsgang  # lokaler Import, zirkulaer sonst

    beruf_namen = {
        b.code: b.name for b in Bildungsgang.query.all() if b.code
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anmeldungen"

    headers = [
        "Nachname", "Vorname", "Geburtsdatum", "Beruf", "Zug",
        "Straße", "PLZ", "Ort", "Telefon", "E-Mail",
        "Eltern/Ansprechpartner", "Eltern-Telefon",
        "Status", "PLZ-Prüfung", "Eingegangen am",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in regs:
        ws.append([
            r.nachname, r.vorname,
            r.geburtsdatum.strftime("%d.%m.%Y") if r.geburtsdatum else None,
            beruf_namen.get(r.beruf, r.beruf), r.zug.name if r.zug else None,
            r.strasse, r.plz, r.ort, r.telefon, r.email,
            f"{r.eltern_vorname or ''} {r.eltern_nachname or ''}".strip() or None,
            r.eltern_telefon,
            r.status, PLZ_OK_LABEL[r.plz_ok],
            r.created_at.strftime("%d.%m.%Y %H:%M") if r.created_at else None,
        ])

    _autosize_columns(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Klassenbuch-Import (schreibt in die admin-hochgeladene Vorlage)
# ---------------------------------------------------------------------------

_KB_DATE_FORMAT = "DD.MM.YYYY"
_KB_DATE_COLUMNS = (4, 6, 7)  # Geburtsdatum, erster Schultag, letzter Schultag


def letzter_schultag(ausbildung_bis: Optional[date]) -> Optional[date]:
    """Letzter Schultag = Ende des Schulhalbjahres, in dem die Ausbildung
    endet: der naechste 31.01. bzw. 31.07. am oder nach `ausbildung_bis`.
    Ohne Ausbildungsende -> None.
    """
    if not ausbildung_bis:
        return None
    y = ausbildung_bis.year
    for kandidat in (date(y, 1, 31), date(y, 7, 31), date(y + 1, 1, 31)):
        if kandidat >= ausbildung_bis:
            return kandidat
    return date(y + 1, 7, 31)


def build_klassenbuch_excel(
    regs, erster_schultag: Optional[date],
) -> io.BytesIO:
    """Wirft FileNotFoundError, wenn keine Vorlage hochgeladen wurde.

    "erster Schultag" ist ein einheitliches Datum fuer den ganzen Export
    (aus den "Gemeinsamen Daten" der Klasse). "letzter Schultag" wird pro
    SuS aus dem Ausbildungsende (`ausbildung_bis`) berechnet - fehlt das,
    bleibt die Zelle leer.

    Die Vorlage bringt uneinheitliche Zell-Formate mit (mal dd/mm/yyyy, mal
    mm-dd-yy, lila Schrift in Zeile 2, Excel-"Tabelle" nur ueber A1:H2).
    Deshalb werden die geschriebenen Datenzeilen hier bewusst normalisiert:
    einheitliches deutsches Datumsformat, Standard-Schrift, und die
    Tabellen-Range wird auf alle Zeilen ausgedehnt (sonst greift die
    Zeilen-Streifung/Faerbung nur auf die erste Zeile).
    """
    path = vorlagen.vorlage_path("klassenbuch")
    if not path.exists():
        raise FileNotFoundError("Klassenbuch-Vorlage wurde noch nicht hochgeladen.")

    wb = openpyxl.load_workbook(path)
    ws = wb["Schüler"]

    first_data_row = 2  # Zeile 1 = Kopfzeile der Vorlage
    row = first_data_row
    klasse_cache = {}
    for r in regs:
        eff_klasse = _effektive_klasse(r, klasse_cache)
        ws.cell(row=row, column=1, value=r.nachname)
        ws.cell(row=row, column=2, value=r.vorname)
        ws.cell(row=row, column=3, value=eff_klasse.name if eff_klasse else None)
        ws.cell(row=row, column=4, value=r.geburtsdatum)
        ws.cell(row=row, column=5, value=GESCHLECHT_KURZFORM.get(r.geschlecht, ""))
        ws.cell(row=row, column=6, value=erster_schultag)
        ws.cell(row=row, column=7, value=letzter_schultag(r.ausbildung_bis))
        ws.cell(row=row, column=8, value=r.email)

        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = Font()  # Vorlagen-Eigenheiten (lila Schrift) entfernen
            if col in _KB_DATE_COLUMNS:
                cell.number_format = _KB_DATE_FORMAT
        row += 1

    last_data_row = row - 1
    if last_data_row >= first_data_row:
        for table in ws.tables.values():
            table.ref = f"A1:H{last_data_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# DaZ-Import (schreibt in die admin-hochgeladene Vorlage, nur daz_bedarf=True)
# ---------------------------------------------------------------------------

def aktuelles_schuljahr(heute: Optional[date] = None) -> str:
    """'2026/27' - Schuljahr beginnt im August. Vor August gehoert das
    aktuelle Kalenderjahr noch zum vorigen Schuljahr.
    """
    heute = heute or date.today()
    start = heute.year if heute.month >= 8 else heute.year - 1
    return f"{start}/{str(start + 1)[-2:]}"


def build_daz_excel(regs) -> io.BytesIO:
    """Wirft FileNotFoundError, wenn keine Vorlage hochgeladen wurde.

    Abteilung wird automatisch aus der admin-gepflegten Zuordnung
    Bildungsgang -> Abteilung ermittelt. Schuljahr aus dem Datum.
    """
    from .models import Bildungsgang  # lokaler Import, zirkulaer sonst

    path = vorlagen.vorlage_path("daz")
    if not path.exists():
        raise FileNotFoundError("DaZ-Vorlage wurde noch nicht hochgeladen.")

    abteilung = None
    for r in regs:
        if r.beruf:
            eintrag = Bildungsgang.query.filter_by(code=r.beruf).first()
            if eintrag and eintrag.abteilung:
                abteilung = eintrag.abteilung.name
                break

    wb = openpyxl.load_workbook(path)
    ws = wb["Tabelle1"]

    ws["B1"] = abteilung
    ws["B4"] = aktuelles_schuljahr()

    row = 13  # Zeile 12 = Kopfzeile der Vorlage
    klasse_cache = {}
    for r in regs:
        if not r.daz_bedarf:
            continue
        eff_klasse = _effektive_klasse(r, klasse_cache)
        ws.cell(row=row, column=1, value=eff_klasse.name if eff_klasse else None)
        ws.cell(row=row, column=2, value=f"{r.nachname or ''}, {r.vorname or ''}".strip(", "))
        ws.cell(row=row, column=3, value=r.ort)
        # Sprachniveau: Spalte D (mit Nachweis) oder E (ohne Nachweis), je
        # nachdem was die LK im Tool eingetragen hat. Fehlt beides, bleibt
        # die Zeile dort leer - zum manuellen Nachtragen.
        if r.sprachniveau:
            col = 4 if r.sprachniveau_nachweis else 5
            ws.cell(row=row, column=col, value=r.sprachniveau)
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Stammdatenblatt-PDF (Koordinaten-Overlay auf die admin-hochgeladene Vorlage)
# ---------------------------------------------------------------------------

def _fmt_date(d) -> Optional[str]:
    return d.strftime("%d.%m.%Y") if d else None


def _combine(*parts) -> Optional[str]:
    joined = " ".join(p for p in parts if p)
    return joined or None


def _ist_zug(klasse) -> bool:
    """Klasse mit Buchstaben-Suffix (ELI026a) = Zug, ohne (ELI026) = die
    unaufgeteilte Klasse eines einzuegigen Bildungsgangs.
    """
    return bool(klasse) and klasse.name[-1:].isalpha()


def _basis_klasse_fuer_beruf(beruf, _cache=None):
    """Die unaufgeteilte Klasse (ohne a/b/c-Suffix) eines einzuegigen
    Bildungsgangs - fuer Anmeldungen, die keinem Zug zugeordnet sind. Nur
    eindeutig, wenn genau eine solche Klasse den Bildungsgang abdeckt.
    """
    if not beruf:
        return None
    if _cache is not None and beruf in _cache:
        return _cache[beruf]
    from .models import Klasse, KlasseBildungsgang  # lokal, zirkulaer sonst
    treffer = [
        k for k in Klasse.query.join(KlasseBildungsgang)
        .filter(KlasseBildungsgang.bildungsgang == beruf).all()
        if not _ist_zug(k)
    ]
    ergebnis = treffer[0] if len(treffer) == 1 else None
    if _cache is not None:
        _cache[beruf] = ergebnis
    return ergebnis


def _effektive_klasse(r, _cache=None):
    """Zug der Anmeldung oder - wenn keiner zugeordnet - die unaufgeteilte
    Klasse des (einzuegigen) Bildungsgangs.
    """
    return r.zug or _basis_klasse_fuer_beruf(r.beruf, _cache)


def _field_values(r, eff_klasse=None) -> dict:
    return {
        "hauptlistennummer": r.hauptlistennummer,
        "aufnahmedatum": _fmt_date(r.aufnahmedatum),
        "zug_name": eff_klasse.name if eff_klasse else None,
        "zug_lehrkraft": eff_klasse.zustaendige_lehrkraft if eff_klasse else None,
        # Eintrittsdatum: primaer pro Anmeldung, sonst die "Gemeinsamen
        # Daten" der Klasse/des Zugs.
        "eintrittsdatum": _fmt_date(
            r.eintrittsdatum or (eff_klasse.eintrittsdatum if eff_klasse else None)
        ),
        "nachname": r.nachname,
        "geburtsland": r.geburtsland,
        "vorname": r.vorname,
        "geburtsort": r.geburtsort,
        "geburtsname": r.geburtsname,
        "staatsangehoerigkeit_1": r.staatsangehoerigkeit_1,
        "staatsangehoerigkeit_2": r.staatsangehoerigkeit_2,
        "geburtsdatum": _fmt_date(r.geburtsdatum),
        "muttersprache": r.muttersprache,
        "jahr_des_zuzugs": r.jahr_des_zuzugs,
        "wohnt_bei": r.wohnt_bei,
        "strasse": r.strasse,
        # Zelle unter "Foerderschwerpunkt:" / rechts von "Art:" - primaer der
        # Foerderschwerpunkt, sonst die (Frei-)Angabe foerderbedarf_art.
        "foerderschwerpunkt_art": r.foerderschwerpunkt or r.foerderbedarf_art,
        "plz_ort": _combine(r.plz, r.ort),
        "kreis": r.kreis,
        "telefon": r.telefon,
        "email": r.email,
        "eltern_nachname": r.eltern_nachname,
        "eltern_vorname": r.eltern_vorname,
        "eltern_strasse": r.eltern_strasse,
        "eltern_plz_ort": _combine(r.eltern_plz, r.eltern_ort),
        "eltern_telefon": r.eltern_telefon,
        "eltern_telefax": r.eltern_telefax,
        "bundesland_kuerzel": r.bundesland_kuerzel,
        "betrieb_name": r.betrieb_name,
        "ausbildung_von": _fmt_date(r.ausbildung_von),
        "ausbildung_bis": _fmt_date(r.ausbildung_bis),
        "betrieb_kreis": r.betrieb_kreis,
        "betrieb_plz_ort": _combine(r.betrieb_plz, r.betrieb_ort),
        "betrieb_strasse": r.betrieb_strasse,
        "betrieb_telefon": r.betrieb_telefon,
        "betrieb_fax": r.betrieb_fax,
        "betrieb_email": r.betrieb_email,
        "beruf": r.beruf,
        "fachrichtung": r.fachrichtung,
        "letzte_schule_kurzform": r.letzte_schule_kurzform,
        "jahr_verlassen": r.jahr_verlassen,
        "klassenstufe": r.klassenstufe,
        "allgemeinbildender_abschluss": r.allgemeinbildender_abschluss,
        "foerderzentrum": r.foerderzentrum,
        "zeugnis_geprueft_am": _fmt_date(r.zeugnis_geprueft_am),
        "bos_fos_beruf": r.bos_fos_beruf,
        "qualifizierungsnachweise_am": _fmt_date(r.qualifizierungsnachweise_am),
        # Rohwerte fuer Checkbox-Vergleich:
        "geschlecht": r.geschlecht,
        "daz_bedarf": r.daz_bedarf,
        "konfession": r.konfession,
        "foerderbedarf": r.foerderbedarf,
        "eltern_ist_vater": r.eltern_ist_vater,
        "eltern_ist_mutter": r.eltern_ist_mutter,
        "eltern_ist_ansprechpartner": r.eltern_ist_ansprechpartner,
        "eltern_hauptwohnsitz": r.eltern_hauptwohnsitz,
        "betrieb_kammer": r.betrieb_kammer,
        "praktikant": r.praktikant,
        "umschueler": r.umschueler,
        "umschulungsvertrag_vorhanden": r.umschulungsvertrag_vorhanden,
        "kostenuebernahme_vorhanden": r.kostenuebernahme_vorhanden,
        "mit_abschluss_beendet": r.mit_abschluss_beendet,
        "art_abschluss_letzte_schule": r.art_abschluss_letzte_schule,
        "lrs": r.lrs,
        "esa_5_jahre_englisch": r.esa_5_jahre_englisch,
        "esa_englisch_ausreichend": r.esa_englisch_ausreichend,
        "zweite_fremdsprache": r.zweite_fremdsprache,
    }


def _build_overlay(r, klasse_cache=None) -> io.BytesIO:
    """Erzeugt ein 2-seitiges PDF nur mit dem Text-/Kreuzchen-Overlay fuer
    eine Registrierung, spaeter per merge_page auf die Vorlage gelegt.
    """
    values = _field_values(r, _effektive_klasse(r, klasse_cache))
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(PAGE_WIDTH, PAGE_HEIGHT))

    for seite in (1, 2):
        for field_page, x, y, attr, fontsize in TEXT_FIELDS:
            if field_page != seite:
                continue
            value = values.get(attr)
            if value:
                c.setFont("Helvetica", fontsize)
                c.drawString(x, PAGE_HEIGHT - y + 1.5, str(value))
        for entry in CHECKBOX_FIELDS:
            field_page, x, y, attr, expected, *rest = entry
            if field_page != seite:
                continue
            if values.get(attr) == expected:
                # optionale 6. Spalte: kleinere Schrift fuer enge Kaestchen
                c.setFont("Helvetica-Bold", rest[0] if rest else 9)
                c.drawString(x, PAGE_HEIGHT - y + 1.5, "X")
        c.showPage()

    c.save()
    buf.seek(0)
    return buf


def build_stammdatenblatt_pdf(regs) -> io.BytesIO:
    """Wirft FileNotFoundError, wenn keine Vorlage hochgeladen wurde.
    Ein PDF mit 2 Seiten je Registrierung, in der Reihenfolge von `regs`.
    """
    path = vorlagen.vorlage_path("stammdatenblatt")
    if not path.exists():
        raise FileNotFoundError("Stammdatenblatt-Vorlage wurde noch nicht hochgeladen.")

    if len(PdfReader(str(path)).pages) < 2:
        raise ValueError("Stammdatenblatt-Vorlage hat weniger als 2 Seiten.")

    writer = PdfWriter()
    klasse_cache = {}
    for r in regs:
        # Frischer Reader je Registrierung: pypdf haelt beim wiederholten
        # append() desselben Reader-Objekts intern dieselben geklonten
        # PageObjects vor, wodurch sich die Overlays mehrerer SuS auf
        # denselben Seiten summiert haetten statt getrennt zu bleiben.
        reader = PdfReader(str(path))
        writer.append(reader)
        idx = len(writer.pages) - 2
        overlay = PdfReader(_build_overlay(r, klasse_cache))
        writer.pages[idx].merge_page(overlay.pages[0])
        writer.pages[idx + 1].merge_page(overlay.pages[1])

    buf = io.BytesIO()
    writer.write(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Namensschilder / Tischkarten (A4, vertikale Faltmarke - zum Aufstellen)
# ---------------------------------------------------------------------------

_NS_MARGIN = 34           # Rand der Blatthaelfte (pt)
_NS_LOGO_MAX_H = 60       # max. Logo-Hoehe (pt)
_NS_NAME_START = 54       # Start-Schriftgroesse fuer den Namen
_NS_NAME_MIN = 14


def _ns_fit_size(lines, max_width) -> int:
    """Groesste Helvetica-Bold-Groesse, bei der jede Zeile in max_width passt."""
    size = _NS_NAME_START
    while size > _NS_NAME_MIN and any(
        stringWidth(t, "Helvetica-Bold", size) > max_width for t in lines
    ):
        size -= 1
    return size


def _ns_draw_rechte_haelfte(c, logo, vorname, nachname):
    """Rechte Blatthaelfte, Inhalt um 90 Grad gegen den Uhrzeigersinn
    gedreht: Schullogo oben links, darunter der Name fett - so steht es
    richtig, wenn das an der senkrechten Faltmarke gefaltete Blatt
    aufgestellt wird. Linke Haelfte bleibt leer.
    """
    half_w = PAGE_WIDTH / 2
    c.saveState()
    # Ursprung in die Mitte der rechten Haelfte, dann 90 Grad CCW.
    c.translate(half_w + half_w / 2, PAGE_HEIGHT / 2)
    c.rotate(90)
    # Lokales Koordinatensystem der aufgestellten Schild-Flaeche:
    #   lokale x-Achse  = Leserichtung  (Blatt-Hoehe,  842 pt)
    #   lokale y-Achse  = "nach oben"   (halbe Blattbreite, ~298 pt)
    face_w, face_h = PAGE_HEIGHT, half_w
    left = -face_w / 2 + _NS_MARGIN
    top = face_h / 2 - _NS_MARGIN

    logo_h = 0.0
    if logo is not None:
        iw, ih = logo.getSize()
        h, w = _NS_LOGO_MAX_H, _NS_LOGO_MAX_H * iw / ih
        max_w = face_w * 0.4
        if w > max_w:
            w, h = max_w, max_w * ih / iw
        c.drawImage(logo, left, top - h, width=w, height=h,
                    mask="auto", preserveAspectRatio=True)
        logo_h = h + 16

    lines = [t for t in (vorname, nachname) if t] or ["—"]
    size = _ns_fit_size(lines, face_w - 2 * _NS_MARGIN)
    lh = size * 1.24
    # Name-Block unter dem Logo, im restlichen Platz vertikal zentriert.
    bereich_oben = top - logo_h
    bereich_unten = -face_h / 2 + _NS_MARGIN
    block_h = lh * len(lines)
    y = (bereich_oben + bereich_unten) / 2 + block_h / 2 - size
    c.setFont("Helvetica-Bold", size)
    c.setFillColorRGB(0, 0, 0)
    for t in lines:
        c.drawCentredString(0, y, t)
        y -= lh
    c.restoreState()


def build_namensschilder_pdf(regs) -> io.BytesIO:
    """Ein A4-Blatt (Hochformat) je Anmeldung mit senkrechter Faltmarke in
    der Mitte. Der Inhalt (Schullogo + Name, um 90 Grad gegen den
    Uhrzeigersinn gedreht) steht auf der rechten Haelfte; die linke bleibt
    leer. Gefaltet und aufgestellt ergibt das ein Namensschild.
    """
    half_w = PAGE_WIDTH / 2
    logo = ImageReader(str(branding.logo_path())) if branding.has_logo() else None

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(PAGE_WIDTH, PAGE_HEIGHT))
    for r in regs:
        # Faltmarke: gestrichelt ueber die volle Hoehe + Ticks an den Kanten.
        c.saveState()
        c.setStrokeColorRGB(0.6, 0.6, 0.6)
        c.setLineWidth(0.7)
        c.setDash(3, 3)
        c.line(half_w, 18, half_w, PAGE_HEIGHT - 18)
        c.setDash()
        c.line(half_w, 0, half_w, 16)
        c.line(half_w, PAGE_HEIGHT, half_w, PAGE_HEIGHT - 16)
        c.setFont("Helvetica", 6)
        c.setFillColorRGB(0.55, 0.55, 0.55)
        c.translate(half_w - 3, 24)
        c.rotate(90)
        c.drawString(0, 0, "hier falten")
        c.restoreState()

        _ns_draw_rechte_haelfte(
            c, logo, (r.vorname or "").strip(), (r.nachname or "").strip()
        )
        c.showPage()

    c.save()
    buf.seek(0)
    return buf
